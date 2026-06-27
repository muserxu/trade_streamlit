"""Momentum Blend Strategy — locked-in design.

This is the vetted strategy arrived at by testing both the asset universe and the
strategy mechanics (see explore_*.py / stress_test_*.py).

  Universe : VOO, QQQ, SMH, XBI, GLD, TLT, DBC   (+ BIL as the cash sleeve)
  Rebalance: monthly, on the last trading day; new holdings take effect the
             next trading day (signal known only at the close).
  Signal   : BLEND of 3/6/12-month total-return momentum (simple average).
  Hold     : top-2 by blended momentum, equal weight (50/50).
  Filter   : DUAL momentum — any pick whose blended momentum is <= BIL's is
             replaced by BIL (go to cash when nothing is trending above cash).

Outputs (in this folder):
  momentum_blend_overall.csv / .xlsx  — daily tracker w/ running YTD_Return
  yearly_returns_blend/yearly_return_blend.csv — per-year summary vs VOO

Run:  python momentum_blend_strategy.py
"""
import os
import pandas as pd
import yfinance as yf

# ============================================================================
# CONFIGURATION  — change the strategy here
# ============================================================================
UNIVERSE = ['VOO', 'QQQ', 'SMH', 'XBI', 'GLD', 'TLT', 'DBC']
CASH = 'BIL'                       # cash / T-bill sleeve for the dual filter
LOOKBACK_MONTHS = [3, 6, 12]       # blended momentum windows
TOP_N = 2                          # number of holdings
USE_DUAL_MOMENTUM = True           # park in CASH when momentum <= cash
START_DATE = '2018-01-01'
BENCHMARK = 'VOO'
TRADING_DAYS_PER_MONTH = 21
# ============================================================================

ALL_TICKERS = UNIVERSE + [CASH]
DAY_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday',
             'Saturday', 'Sunday']


# ----------------------------------------------------------------------------
# Data
# ----------------------------------------------------------------------------
def fetch_prices(symbols, start=START_DATE):
    """Fetch daily closes for all symbols into one aligned DataFrame."""
    data = {}
    for sym in symbols:
        print(f"Fetching {sym}...")
        hist = yf.Ticker(sym).history(start=start)
        if hist.empty:
            raise RuntimeError(f"No data returned for {sym}")
        s = hist['Close']
        if s.index.tz is not None:
            s.index = s.index.tz_localize(None)
        data[sym] = s
    df = pd.DataFrame(data).sort_index()
    df = df.dropna(how='all')
    return df


# ----------------------------------------------------------------------------
# Momentum signal
# ----------------------------------------------------------------------------
def _trailing_return(prices, asset, end_idx, lookback_days):
    start_idx = end_idx - lookback_days
    if start_idx < 0:
        return None
    p0, p1 = prices[asset].iloc[start_idx], prices[asset].iloc[end_idx]
    if pd.isna(p0) or pd.isna(p1) or p0 == 0:
        return None
    return p1 / p0 - 1


def blended_momentum(prices, asset, end_idx):
    """Average total-return momentum across LOOKBACK_MONTHS, as % ."""
    vals = []
    for m in LOOKBACK_MONTHS:
        r = _trailing_return(prices, asset, end_idx, m * TRADING_DAYS_PER_MONTH)
        if r is not None:
            vals.append(r)
    if not vals:
        return None
    return sum(vals) / len(vals) * 100.0


def month_end_indices(prices):
    """Integer positions of the last trading day of each month."""
    idx = prices.index
    out = []
    for i in range(len(idx)):
        is_last = (i == len(idx) - 1) or (idx[i + 1].month != idx[i].month) \
            or (idx[i + 1].year != idx[i].year)
        if is_last:
            out.append(i)
    return out


# ----------------------------------------------------------------------------
# Strategy: decide month-end holdings, then expand to daily weights
# ----------------------------------------------------------------------------
def decide_holdings(prices, end_idx):
    """Return (weights dict, scores dict) for the decision at end_idx."""
    scores = {}
    for a in UNIVERSE:
        s = blended_momentum(prices, a, end_idx)
        if s is not None:
            scores[a] = s
    cash_score = blended_momentum(prices, CASH, end_idx)
    cash_score = 0.0 if cash_score is None else cash_score

    weights = {}
    if scores:
        ranked = sorted(scores, key=scores.get, reverse=True)
        picks = []
        for a in ranked:
            if len(picks) >= TOP_N:
                break
            if USE_DUAL_MOMENTUM and scores[a] <= cash_score:
                continue          # fails absolute filter -> slot goes to cash
            picks.append(a)
        w = 1.0 / TOP_N
        for a in picks:
            weights[a] = w
        cash_w = 1.0 - sum(weights.values())
        if cash_w > 1e-9:
            weights[CASH] = weights.get(CASH, 0.0) + cash_w
    return weights, scores


def build_daily_frame(prices):
    """Build a per-trading-day DataFrame with holdings, allocation, signal,
    daily strategy return, and a running YTD return that resets each year."""
    me = month_end_indices(prices)
    me_set = set(me)
    daily_ret = prices.pct_change()

    # Decision made at each month-end takes effect from the NEXT trading day.
    decisions = {}   # effective_from_index -> (weights, scores, decided_on_date)
    for ei in me:
        weights, scores = decide_holdings(prices, ei)
        eff = ei + 1
        if eff < len(prices):
            decisions[eff] = (weights, scores, prices.index[ei])

    rows = []
    cur_weights = {}
    cur_decided_on = None
    cur_year = None
    ytd_growth = 1.0
    prev_weights_for_change = {}

    for i in range(len(prices)):
        date = prices.index[i]

        # apply a decision that becomes effective today
        rebalanced = False
        if i in decisions:
            new_w, _scores, decided_on = decisions[i]
            if new_w != cur_weights:
                rebalanced = True
            prev_weights_for_change = cur_weights
            cur_weights = new_w
            cur_decided_on = decided_on

        # daily portfolio return from current weights
        day_ret = 0.0
        for a, w in cur_weights.items():
            dr = daily_ret[a].iloc[i]
            if pd.notna(dr):
                day_ret += w * dr

        # YTD reset at the start of each calendar year
        if date.year != cur_year:
            cur_year = date.year
            ytd_growth = 1.0
        ytd_growth *= (1 + day_ret)
        ytd_pct = (ytd_growth - 1) * 100

        row = {
            'Date': date.strftime('%Y-%m-%d'),
            'Day_of_Week': DAY_NAMES[date.dayofweek],
        }
        # Each ticker's price immediately followed by its blended-momentum SIGNAL
        # (the score that drives the picks). Pairing price+mom side by side makes
        # it easy to watch one asset overtake another. BIL_mom is the cash
        # threshold for the dual-momentum filter.
        day_scores = {}
        for a in ALL_TICKERS:   # include BIL (cash sleeve) so every holdable asset has columns
            row[a] = prices[a].iloc[i]
            s = blended_momentum(prices, a, i)
            day_scores[a] = s
            row[f'{a}_mom'] = round(s, 2) if s is not None else None
        # Which two the signal currently favors (ignores the monthly lag) and the
        # current leader, so you can see the signal vs. what is actually held.
        ranked_now = sorted((a for a in UNIVERSE if day_scores[a] is not None),
                            key=lambda a: day_scores[a], reverse=True)
        row['Signal_Top2'] = ' / '.join(ranked_now[:2]) if ranked_now else 'N/A'

        # current allocation string
        row['Allocation'] = format_allocation(cur_weights)
        # rebalance / decision marker (month-end signal that took effect today)
        row['Rebalance'] = describe_rebalance(prev_weights_for_change, cur_weights) \
            if rebalanced else ''
        row['Day_Return_Pct'] = day_ret * 100
        row['YTD_Return'] = f"{ytd_pct:+.2f}%"
        rows.append(row)

    return pd.DataFrame(rows)


def format_allocation(weights):
    if not weights:
        return 'N/A'
    parts = [f"{a} ({w*100:.0f}%)" for a, w in
             sorted(weights.items(), key=lambda kv: -kv[1])]
    return ' / '.join(parts)


def describe_rebalance(prev, new):
    """Short description of what changed at a rebalance."""
    added = [a for a in new if a not in prev]
    removed = [a for a in prev if a not in new]
    bits = []
    if added:
        bits.append('BUY ' + ', '.join(added))
    if removed:
        bits.append('SELL ' + ', '.join(removed))
    return '; '.join(bits) if bits else 'rebalance'


# ----------------------------------------------------------------------------
# Yearly summary vs VOO buy & hold
# ----------------------------------------------------------------------------
def yearly_summary(daily_df, prices):
    df = daily_df.copy()
    df['Date'] = pd.to_datetime(df['Date'])
    df['Year'] = df['Date'].dt.year
    df['strat_factor'] = 1 + df['Day_Return_Pct'] / 100

    voo_ret = prices[BENCHMARK].pct_change().fillna(0)
    voo_by_date = dict(zip(prices.index, voo_ret.values))
    df['voo_factor'] = df['Date'].map(lambda d: 1 + voo_by_date.get(d, 0.0))

    out = []
    strat_cum, voo_cum = 1.0, 1.0
    for year in sorted(df['Year'].unique()):
        yd = df[df['Year'] == year]
        sg = yd['strat_factor'].prod()
        vg = yd['voo_factor'].prod()
        strat_cum *= sg
        voo_cum *= vg
        out.append({
            'Year': year,
            'Strategy_Pct': (sg - 1) * 100,
            'VOO_Pct': (vg - 1) * 100,
            'Outperf_Pct': (sg - vg) * 100,
            'Strat_Cumulative_Pct': (strat_cum - 1) * 100,
            'VOO_Cumulative_Pct': (voo_cum - 1) * 100,
        })
    return pd.DataFrame(out)


def print_yearly(summary_df):
    print(f"\n{'='*78}")
    print("  MOMENTUM BLEND STRATEGY — Yearly Returns vs VOO Buy & Hold")
    print(f"{'='*78}")
    print(f"  {'Year':<8}{'Strategy':>12}{'VOO B&H':>12}{'Diff':>12}{'Strat Cum':>14}")
    print(f"  {'-'*8}{'-'*12}{'-'*12}{'-'*12}{'-'*14}")
    for _, r in summary_df.iterrows():
        yr = int(r['Year'])
        label = f"{yr} YTD" if yr == 2026 else str(yr)
        print(f"  {label:<8}{r['Strategy_Pct']:>+11.2f}%{r['VOO_Pct']:>+11.2f}%"
              f"{r['Outperf_Pct']:>+11.2f}%{r['Strat_Cumulative_Pct']:>+13.2f}%")
    final = summary_df.iloc[-1]
    print(f"  {'-'*8}{'-'*12}{'-'*12}{'-'*12}{'-'*14}")
    print(f"  {'TOTAL':<8}{'':>12}{'':>12}{'':>12}"
          f"{final['Strat_Cumulative_Pct']:>+13.2f}%")


# ----------------------------------------------------------------------------
# Save: CSV + formatted Excel
# ----------------------------------------------------------------------------
def save_outputs(daily_df, summary_df):
    # Display newest day at the top, oldest at the bottom (computation above was
    # chronological, so YTD/momentum are already correct — we only reverse rows).
    daily_df = daily_df.iloc[::-1].reset_index(drop=True)

    daily_df.to_csv('momentum_blend_overall.csv', index=False)
    print("\nSaved momentum_blend_overall.csv")

    out_dir = 'yearly_returns_blend'
    os.makedirs(out_dir, exist_ok=True)
    summary_df.to_csv(os.path.join(out_dir, 'yearly_return_blend.csv'), index=False)
    print(f"Saved {out_dir}/yearly_return_blend.csv")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        xlsx = 'momentum_blend_overall.xlsx'
        with pd.ExcelWriter(xlsx, engine='openpyxl') as writer:
            daily_df.to_excel(writer, sheet_name='Daily', index=False)
            summary_df.to_excel(writer, sheet_name='Yearly', index=False)

        wb = load_workbook(xlsx)
        header_fill = PatternFill('solid', fgColor='366092')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        date_fill = PatternFill('solid', fgColor='D9E1F2')
        day_fill = PatternFill('solid', fgColor='FFF2CC')
        alloc_fill = PatternFill('solid', fgColor='E2EFDA')
        rebal_fill = PatternFill('solid', fgColor='FFE699')
        ytd_pos = PatternFill('solid', fgColor='C6EFCE')
        ytd_neg = PatternFill('solid', fgColor='FFC7CE')
        mom_fill = PatternFill('solid', fgColor='F2F2F2')        # momentum signal block
        leader_fill = PatternFill('solid', fgColor='FFC7CE')     # top-2 momentum pick (red = "the chosen")
        signal_fill = PatternFill('solid', fgColor='DDEBF7')     # Signal_Top2 column

        ws = wb['Daily']
        col_by_hdr = {c.value: c.column for c in ws[1]}
        # momentum columns for the rotation assets (exclude BIL — it's the threshold)
        mom_cols = {a: col_by_hdr.get(f'{a}_mom') for a in UNIVERSE}
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Pre-pass: find the two highest momentum cells in this row so we can
            # highlight them (this is the strategy's actual pick logic, visualized).
            row_idx = row[0].row
            scored = []
            for a, ci in mom_cols.items():
                if ci is None:
                    continue
                v = ws.cell(row=row_idx, column=ci).value
                if isinstance(v, (int, float)):
                    scored.append((v, ci))
            top2_cols = {ci for _, ci in sorted(scored, reverse=True)[:2]}

            for cell in row:
                hdr = ws.cell(row=1, column=cell.column).value
                if hdr == 'Date':
                    cell.fill = date_fill
                elif hdr == 'Day_of_Week':
                    cell.fill = day_fill
                    cell.alignment = Alignment(horizontal='center')
                elif hdr == 'Signal_Top2':
                    cell.fill = signal_fill
                    cell.alignment = Alignment(horizontal='center')
                elif hdr == 'Allocation':
                    cell.fill = alloc_fill
                    cell.alignment = Alignment(horizontal='center')
                elif hdr == 'Rebalance':
                    cell.alignment = Alignment(horizontal='center')
                    if cell.value:
                        cell.fill = rebal_fill
                        cell.font = Font(bold=True)
                elif hdr == 'YTD_Return':
                    cell.alignment = Alignment(horizontal='center')
                    v = str(cell.value)
                    if v.startswith('-'):
                        cell.fill = ytd_neg
                        cell.font = Font(color='9C0006', bold=True)
                    elif v not in ('', 'None'):
                        cell.fill = ytd_pos
                        cell.font = Font(color='006100', bold=True)
                elif hdr and str(hdr).endswith('_mom'):
                    # momentum signal: format, color sign, highlight the top-2 picks
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '+0.0;-0.0'
                        if cell.column in top2_cols:
                            cell.fill = leader_fill          # this asset is a current pick
                            cell.font = Font(bold=True, color='9C0006')
                        else:
                            cell.fill = mom_fill
                            cell.font = Font(color='006100' if cell.value >= 0 else 'C00000')
                    else:
                        cell.fill = mom_fill
                elif hdr in ALL_TICKERS or hdr == 'Day_Return_Pct':
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'

        # Yearly sheet formatting
        ws2 = wb['Yearly']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            for cell in row:
                hdr = ws2.cell(row=1, column=cell.column).value
                if hdr != 'Year' and isinstance(cell.value, (int, float)):
                    cell.number_format = '+0.00"%";-0.00"%"'
                    if cell.value > 0:
                        cell.fill = ytd_pos
                        cell.font = Font(color='006100')
                    elif cell.value < 0:
                        cell.fill = ytd_neg
                        cell.font = Font(color='9C0006')

        for sheet in (ws, ws2):
            for col in sheet.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None),
                            default=8)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = \
                    min(width + 3, 40)
            sheet.freeze_panes = 'A2'

        wb.save(xlsx)
        print(f"Saved {xlsx} (formatted)")
    except ImportError:
        print("openpyxl not installed — skipping Excel output.")
    except Exception as e:
        print(f"Could not write Excel: {e}")


# ----------------------------------------------------------------------------
def main():
    print("=" * 78)
    print("  MOMENTUM BLEND STRATEGY")
    print(f"  Universe : {', '.join(UNIVERSE)}  | cash: {CASH}")
    print(f"  Signal   : blend {'/'.join(map(str, LOOKBACK_MONTHS))}-month momentum, "
          f"top-{TOP_N}, {'dual' if USE_DUAL_MOMENTUM else 'relative'} momentum")
    print(f"  Rebalance: monthly (last trading day), effective next day")
    print("=" * 78)

    prices = fetch_prices(ALL_TICKERS)
    print(f"\nData range: {prices.index.min():%Y-%m-%d} to {prices.index.max():%Y-%m-%d} "
          f"({len(prices)} trading days)")

    daily_df = build_daily_frame(prices)
    summary_df = yearly_summary(daily_df, prices)

    print_yearly(summary_df)

    # Current standing
    last = daily_df.iloc[-1]
    print(f"\n  Current allocation ({last['Date']}): {last['Allocation']}")
    print(f"  2026 YTD return: {last['YTD_Return']}")

    save_outputs(daily_df, summary_df)
    print("\nDone.")


if __name__ == '__main__':
    main()
